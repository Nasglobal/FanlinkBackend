from django.shortcuts import render,get_object_or_404
from django.forms.models import model_to_dict
from django.db.models import Q,Sum
from django.utils import timezone
from django.conf import settings
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import subprocess
from django.db.models.functions import Trim
from PIL import Image
from django.core.files.storage import default_storage
from django.http import JsonResponse,StreamingHttpResponse,HttpResponse,FileResponse,Http404
from .models import MediaFiles,FanLinks,Releases,Video
from rest_framework import viewsets, mixins, status,pagination
from rest_framework.decorators import action,api_view,parser_classes
from .serializers import MediaFilesSerializers,FanlinksSerializers,ReleasesSerializers
from django.views.decorators.csrf import csrf_exempt
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.models import User
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.hashers import check_password
from rest_framework.permissions import IsAuthenticated
from googleapiclient.discovery import build
from google.oauth2 import service_account
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
from .utils import fetch_sheet_data,get_last_updated_row,get_google_credentials
from sendfile import sendfile
from django.urls import reverse
import openpyxl
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,PatternFill, Border as XLBorder, Side
import csv
from io import StringIO,BytesIO,TextIOWrapper
import os
import sys
import shutil
import ffmpeg
import pandas as pd
from google.oauth2.service_account import Credentials
from gspread_formatting import (
    format_cell_range,
    CellFormat,
    Borders,
    Border,
    Color,
    TextFormat,
    NumberFormat,
    batch_updater
)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import ZipFile
import numpy as np
import io
import random
import string
import requests
from datetime import datetime
from .spotifyFunctions import get_spotify_track_link,replace_spaces_with_underscore,search_boomplay_with_google,search_audiomack_with_google,get_itunes_track_link
from .youtubeFunctions import get_youtube_video_link,get_deezer_track_link,get_apple_music_link,search_amazon_music_with_google,search_tidal_with_google

# Global variable to store previously fetched rows
previous_rows = []

class CustomPagination(pagination.PageNumberPagination):
    page_size = 1000  # Number of records per page
    page_size_query_param = 'page_size'
    max_page_size = 2000

    def get_page_size(self, request):
        page_size = request.query_params.get('page_size') or self.page_size
        return min(int(page_size), self.max_page_size)

    def get_paginated_response(self, data):
        total_items = self.page.paginator.count
        return Response({
            'total_items': total_items,
            'page_number': self.page.number,
            'page_size': self.page.paginator.per_page,
            'results': data
        })

# Create your views here.

class MediaFileViewset(viewsets.ModelViewSet):
  queryset = MediaFiles.objects.filter(DataType='Track').order_by('id')
  serializer_class = MediaFilesSerializers
  pagination_class = CustomPagination  # Use the custom pagination class 
 
  def get_queryset(self):
      min_id = self.request.query_params.get('min_id')
      max_id = self.request.query_params.get('max_id')

      queryset = MediaFiles.objects.filter(DataType='Track').order_by('id')

      if min_id:
          queryset = queryset.filter(id__gte=min_id)

      if max_id:
          queryset = queryset.filter(id__lte=max_id)
      
      return queryset

class FanLinksViewSet(viewsets.ModelViewSet):
    queryset = FanLinks.objects.all()
    serializer_class = FanlinksSerializers
    pagination_class = CustomPagination


    @action(detail=False, methods=["POST"]) #convert this to endpoint
    def create_fanlink(self, request, *args, **kwargs):
        artist_name = request.data.get("artist")
        track_name = request.data.get("track")
        description = request.data.get("description")
        release_date = request.data.get("releaseDate")
        isrc = request.data.get("isrc")
        source = request.data.get("source")
        label_name = request.data.get("label")

        artist = replace_spaces_with_underscore(artist_name)
        track = replace_spaces_with_underscore(track_name)
        missingLinks = ""
        
        if not artist_name or not track_name:
            return JsonResponse({"error": "Artist name and track name are required."}, status=400)
        
        track_link = get_spotify_track_link(artist_name, track_name, release_date,isrc)
        video_link = get_youtube_video_link(artist_name, track_name, release_date,isrc)
        boomplay_link = search_boomplay_with_google(artist_name, track_name, release_date,isrc)
        audiomack_link = search_audiomack_with_google(artist_name, track_name, release_date,isrc)
        itunes_link = get_itunes_track_link(artist_name, track_name, release_date,isrc)
        deezer_link = get_deezer_track_link(artist_name, track_name, release_date,isrc)
        apple_music_link = get_apple_music_link(artist_name, track_name, release_date,isrc)
        amazon_music_link = search_amazon_music_with_google(artist_name, track_name, release_date,isrc)
        tidal_link = search_tidal_with_google(artist_name, track_name, release_date,isrc)


        if track_link or video_link or boomplay_link or audiomack_link or itunes_link or deezer_link or apple_music_link or amazon_music_link or tidal_link:
            fanlinks = f"/{artist}-{track}"
            fanlink = f"https://fanlink.51lexapps.com/{artist}-{track}"
            try:
               check_fan_links = FanLinks.objects.get(ArtistName=artist,TrackName=track)
               check_fan_links.SpotifyLink = track_link
               check_fan_links.YoutubeLink = video_link
               check_fan_links.Boomplay = boomplay_link
               check_fan_links.AudiomackLink = audiomack_link
               check_fan_links.ItunesLink = itunes_link
               check_fan_links.DeezerLink = deezer_link
               check_fan_links.AppleLink = apple_music_link
               check_fan_links.AmazonLink = amazon_music_link
               check_fan_links.TidalLink = tidal_link
               check_fan_links.Source = source
               check_fan_links.save()
               
            except FanLinks.DoesNotExist:
                FanLinks(ArtistName=artist,TrackName=track,SpotifyLink=track_link,AppleLink=apple_music_link,AmazonLink=amazon_music_link,YoutubeLink=video_link,ItunesLink=itunes_link,AudiomackLink=audiomack_link,DeezerLink=deezer_link,TidalLink=tidal_link,Boomplay=boomplay_link,Description=description,UPC=isrc,ReleaseDate=release_date,Source=source).save()
            if not track_link : 
                missingLinks = missingLinks + "Spotify,"
            if not video_link:
                missingLinks = missingLinks + "Youtube,"
            if not boomplay_link:
                missingLinks = missingLinks + "Boomplay,"
            if not audiomack_link:
                missingLinks = missingLinks + "Audiomack,"
            if not itunes_link:
                missingLinks = missingLinks + "Itunes,"
            if not deezer_link:
                missingLinks = missingLinks + "Deezer,"
            if not apple_music_link:
                missingLinks = missingLinks + "Apple Music,"
            if not amazon_music_link:
                missingLinks = missingLinks + "Amazon,"
            if not tidal_link:
                missingLinks = missingLinks + "Tidal"
            try:
                releasesList = Releases.objects.filter(Artists=artist_name,Title=track_name).first()
                if releasesList is not None:
                    releasesList.FanlinkSent = fanlink
                    releasesList.MissingLinks=missingLinks
                    releasesList.save()
                else:
                   Releases(Label=label_name,Artists=artist_name,Title=track_name,UPC=isrc,ReleaseDate=release_date,FanlinkSent=fanlink,Status="",Y="",MissingLinks=missingLinks).save() 
            except Releases.DoesNotExist:
                Releases(Label=label_name,Artists=artist_name,Title=track_name,UPC=isrc,ReleaseDate=release_date,FanlinkSent=fanlink,Status="",Y="",MissingLinks=missingLinks).save()
            
            return JsonResponse({"link": fanlinks})
        else:
            return JsonResponse({"error": "Track not found."}, status=404)


class ReleasesViewSet(viewsets.ModelViewSet):
    queryset = Releases.objects.all().order_by('id')
    serializer_class = ReleasesSerializers
    pagination_class = CustomPagination

    @action(detail=False, methods=["POST"]) #convert this to endpoint
    def upload_releases(self, request, *args, **kvargs):
        releases = request.FILES.get('releases')
        csv_data = releases.read()
        new_data_frame = pd.read_excel(csv_data)
        relevant_columns = [
            'Label', 'Artists', 'Title', 
            'UPC', 'ReleaseDate', 'FanlinkSent', 
            'Status', 'Y', 'MissingLinks'
        ]
        new_data_frame = new_data_frame[relevant_columns]
        for _, row in new_data_frame.iterrows():
          artist_name = row['Artists']
          track_name = row['Title']
          

          if Releases.objects.filter(Artists__iexact=artist_name,Title__iexact=track_name).exists():
              print("track name exist already : ", track_name)
              # Skip existing rows
              continue
          else:
              # Replace NaN with empty string
              cleaned_row = row.where(pd.notna(row), '')
              # Convert row to dictionary
              row_dict = cleaned_row.to_dict()
              # Create the new record
              Releases.objects.create(**row_dict)  

        return JsonResponse({'message': "Releases upload successful"})



class RegisterView(APIView):
    def post(self, request):
        username = request.data.get('username')
        email = request.data.get('email')
        password = request.data.get('password')

        if User.objects.filter(username=username).exists():
            return Response({"error": "Username already exists"}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=email).exists():
            return Response({"error": "Email already exists"}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(username=username, email=email, password=password)
        user.save()

        # Generate token for the user
        refresh = RefreshToken.for_user(user)
        return Response({
            "message": "User created successfully",
            "refresh": str(refresh),
            "access": str(refresh.access_token),
        }, status=status.HTTP_201_CREATED)

class LoginView(APIView):
    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')

        try:
            # Find the user by email
            user = User.objects.get(email=email)
            # Verify the hashed password
            if check_password(password, user.password):
                refresh = RefreshToken.for_user(user)
                return Response({
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }, status=status.HTTP_200_OK)
            else:
                return Response({"error": "Invalid password"}, status=status.HTTP_401_UNAUTHORIZED)
        except User.DoesNotExist:
            return Response({"error": "User with this email does not exist"}, status=status.HTTP_404_NOT_FOUND)


class ProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            "username": user.username,
            "email": user.email,
        }, status=status.HTTP_200_OK)


@csrf_exempt
def get_fanlink(request,track,artist):
      try:
        all_fan_links = FanLinks.objects.get(ArtistName=artist,TrackName=track)
        if not all_fan_links:
          return JsonResponse({"error": "No fan link available."}, status=400)    
        else:
         res = {
          "spotifyLink" : all_fan_links.SpotifyLink,
          "appleLink" : all_fan_links.AppleLink,
          "amazonLink" : all_fan_links.AmazonLink,
          "youtubeLink" : all_fan_links.YoutubeLink,
          "itunesLink" : all_fan_links.ItunesLink,
          "audiomackLink" : all_fan_links.AudiomackLink,
          "deezerLink" : all_fan_links.DeezerLink,
          "tidalLink" : all_fan_links.TidalLink,
          "boomplay" : all_fan_links.Boomplay,
          "source" : all_fan_links.Source,
          "upc" : all_fan_links.UPC,
         }
        return JsonResponse({'data': res })
      except FanLinks.DoesNotExist:
        return JsonResponse({"error": "Artist name or track name does not exist."}, status=400)


@csrf_exempt
def drive_webhook(request):
    """Handle incoming webhook notifications from Google Drive."""
    global previous_rows
    spreadsheet_id = "16dMltfMyyl8WAEDy9ZRxu3kLpFYUNb7ktMB5SGSV8EY"
    range_name = "Sheet1!A1:Z1000"
    if request.method == 'POST':
        content_type = request.headers.get('Content-Type')
        if content_type == 'application/json':
            try:
                # Parse the JSON body
                notification = json.loads(request.body)
                print("Webhook triggered with JSON body:", notification)
                current_rows = fetch_sheet_data(spreadsheet_id, range_name)

                # Detect newly added rows
                if not previous_rows:
                    previous_rows = current_rows  # Initialize if empty
                    print("Initialized previous rows.")
                else:
                    new_rows = current_rows[len(previous_rows):]
                    if new_rows:
                        print("Newly added rows:")
                        for row in new_rows:
                            print(row)
                            auto_generate_fanlink(row[1],row[2],row[0],row[6 ],row[4])
                    previous_rows = current_rows
                return JsonResponse({"message": "Notification received successfully."}, status=200)
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid JSON data"}, status=400)
        elif content_type == 'text/plain' and request.body == b'':
            try:
                current_rows = get_last_updated_row(spreadsheet_id, range_name)
                # Detect newly added rows
                if not previous_rows:
                    previous_rows = current_rows  # Initialize if empty
                else:
                    new_rows = current_rows[len(previous_rows):]
                    if new_rows:
                        for row in new_rows:
                            print(row)
                            auto_generate_fanlink(row[1],row[2],row[0],row[6],row[4])

                    # Update the stored rows
                    previous_rows = current_rows

                return JsonResponse(
                    {"message": "Notification processed", "last_row": current_rows},
                    status=200,
                )
            except Exception as e:
                print("Error querying Google Sheets API:", str(e))
                return JsonResponse({"error": str(e)}, status=500)
        
            return JsonResponse({"message": "Notification received with no body."}, status=200)

    return HttpResponse("Invalid request", status=400)

def auto_generate_fanlink(artist_name,track_name,label_name,isrc,release_date): 
    artist = replace_spaces_with_underscore(artist_name)
    track = replace_spaces_with_underscore(track_name)
    missingLinks = ""
    src = ""
    if not artist_name or not track_name:
        print("Artist name and track name are required.")
    else:   
        track_link = get_spotify_track_link(artist_name, track_name, release_date,isrc)
        video_link = get_youtube_video_link(artist_name, track_name, release_date,isrc)
        boomplay_link = search_boomplay_with_google(artist_name, track_name, release_date,isrc)
        audiomack_link = search_audiomack_with_google(artist_name, track_name, release_date,isrc)
        itunes_link = get_itunes_track_link(artist_name, track_name, release_date,isrc)
        deezer_link = get_deezer_track_link(artist_name, track_name, release_date,isrc)
        apple_music_link = get_apple_music_link(artist_name, track_name, release_date,isrc)
        amazon_music_link = search_amazon_music_with_google(artist_name, track_name, release_date,isrc)
        tidal_link = search_tidal_with_google(artist_name, track_name, release_date,isrc)

        if track_link or video_link or boomplay_link or audiomack_link or itunes_link or deezer_link or apple_music_link or amazon_music_link or tidal_link:
            if video_link :
                src = "youtube"
            else:
                src = "spotify"
            try:
               check_fan_links = FanLinks.objects.get(ArtistName=artist,TrackName=track)
               check_fan_links.SpotifyLink = track_link
               check_fan_links.YoutubeLink = video_link
               check_fan_links.Boomplay = boomplay_link
               check_fan_links.AudiomackLink = audiomack_link
               check_fan_links.ItunesLink = itunes_link
               check_fan_links.DeezerLink = deezer_link
               check_fan_links.AppleLink = apple_music_link
               check_fan_links.AmazonLink = amazon_music_link
               check_fan_links.TidalLink = tidal_link
               check_fan_links.Source = src
               check_fan_links.save()
            except FanLinks.DoesNotExist:
                FanLinks(ArtistName=artist,TrackName=track,SpotifyLink=track_link,AppleLink=apple_music_link,AmazonLink=amazon_music_link,YoutubeLink=video_link,ItunesLink=itunes_link,AudiomackLink=audiomack_link,DeezerLink=deezer_link,TidalLink=tidal_link,Boomplay=boomplay_link,Description="auto generated",UPC=isrc,ReleaseDate=release_date,Source=src).save()
            fanlink = f"/{artist}-{track}"
            Releases(Label=label_name,Artists=artist_name,Title=track_name,UPC=isrc,ReleaseDate=release_date,FanlinkSent=fanlink,Status="",Y="",MissingLinks="").save()



def generate_fanlink_toSheet(artist_name,track_name,label_name,isrc,release_date): 
    artist = replace_spaces_with_underscore(artist_name)
    track = replace_spaces_with_underscore(track_name)
    missingLinks = ""
    src = ""
    if not artist_name or not track_name:
        print("Artist name and track name are required.")
    else:   
        track_link = get_spotify_track_link(artist_name, track_name, release_date,isrc)
        video_link = get_youtube_video_link(artist_name, track_name, release_date,isrc)
        boomplay_link = search_boomplay_with_google(artist_name, track_name, release_date,isrc)
        audiomack_link = search_audiomack_with_google(artist_name, track_name, release_date,isrc)
        itunes_link = get_itunes_track_link(artist_name, track_name, release_date,isrc)
        deezer_link = get_deezer_track_link(artist_name, track_name, release_date,isrc)
        apple_music_link = get_apple_music_link(artist_name, track_name, release_date,isrc)
        amazon_music_link = search_amazon_music_with_google(artist_name, track_name, release_date,isrc)
        tidal_link = search_tidal_with_google(artist_name, track_name, release_date,isrc)

        if track_link or video_link or boomplay_link or audiomack_link or itunes_link or deezer_link or apple_music_link or amazon_music_link or tidal_link:
            if video_link :
                src = "youtube"
            else:
                src = "spotify"

            try:
               check_fan_links = FanLinks.objects.get(ArtistName=artist,TrackName=track)
               check_fan_links.SpotifyLink = track_link
               check_fan_links.YoutubeLink = video_link
               check_fan_links.Boomplay = boomplay_link
               check_fan_links.AudiomackLink = audiomack_link
               check_fan_links.ItunesLink = itunes_link
               check_fan_links.DeezerLink = deezer_link
               check_fan_links.AppleLink = apple_music_link
               check_fan_links.AmazonLink = amazon_music_link
               check_fan_links.TidalLink = tidal_link
               check_fan_links.Source = src
               check_fan_links.save()
            except FanLinks.DoesNotExist:
                FanLinks(ArtistName=artist,TrackName=track,SpotifyLink=track_link,AppleLink=apple_music_link,AmazonLink=amazon_music_link,YoutubeLink=video_link,ItunesLink=itunes_link,AudiomackLink=audiomack_link,DeezerLink=deezer_link,TidalLink=tidal_link,Boomplay=boomplay_link,Description="auto generated",UPC=isrc,ReleaseDate=release_date,Source=src).save()
            fanlink = f"https://fanlink.51lexapps.com/{artist}-{track}"
            if not track_link : 
                missingLinks = missingLinks + "Spotify,"
            if not video_link:
                missingLinks = missingLinks + "Youtube,"
            if not boomplay_link:
                missingLinks = missingLinks + "Boomplay,"
            if not audiomack_link:
                missingLinks = missingLinks + "Audiomack,"
            if not itunes_link:
                missingLinks = missingLinks + "Itunes,"
            if not deezer_link:
                missingLinks = missingLinks + "Deezer,"
            if not apple_music_link:
                missingLinks = missingLinks + "Apple Music,"
            if not amazon_music_link:
                missingLinks = missingLinks + "Amazon,"
            if not tidal_link:
                missingLinks = missingLinks + "Tidal"

            Releases(Label=label_name,Artists=artist_name,Title=track_name,UPC=isrc,ReleaseDate=release_date,FanlinkSent=fanlink,Status="",Y="",MissingLinks=missingLinks).save()
            return {"fanlink":fanlink,"missingLinks":missingLinks}
        else:
            return {"fanlink":"Fanlink not found","missingLinks":missingLinks}

@csrf_exempt
def generate_fanlinks_in_batch(request):
    # Authenticate and connect to Google Sheets
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        # Get the absolute path to the root directory
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        # Full path to your credentials file
        creds_path = os.path.join(BASE_DIR, "fanlink-440822-6316459498b3.json")

        # Use the credentials path
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
        client = gspread.authorize(creds)
        # Open Google Sheet by its key
        spreadsheet = client.open_by_key("1j4DSWgEECumRDfJ6ZEOLg4NapMpxgRa-dX6QMifQCy0") 
        # Open the specific worksheet
        sheet = spreadsheet.worksheet("The Orchard")  # Change "Sheet1" if needed
        expected_headers = ['Label', 'Artist', 'Release', 'UPC', 'Date', 'Links','ISRC','Fanlinks','MissingLinks']
        all_data = sheet.get_all_records(expected_headers=expected_headers)
        print("connected to google sheet...")
        empty_fanlink_rows = [i for i, row in enumerate(all_data) if row.get('Fanlinks', '').strip() == '']
        if not empty_fanlink_rows:
            return JsonResponse({"message": "No unprocessed tracks found"})
        else:
            target_indexes = empty_fanlink_rows[:2]

            for idx in target_indexes:
                row_data = all_data[idx]
                #fanlink = generate_fanlink_toSheet(row_data) 
                fanlink = generate_fanlink_toSheet(row_data["Artist"],row_data["Release"],row_data["Label"],row_data["ISRC"],row_data["Date"])
                sheet.update_cell(idx + 2, 8, fanlink["fanlink"])  # Row index +2 (header + 1-based)
                sheet.update_cell(idx + 2, 9, fanlink["missingLinks"])
            return JsonResponse({"message": "Fanlinks updated for rows"})
    except Exception as e:  
        #-7th dec pull back      
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def export_releases_fanlink(request):
    # Generate Excel file
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    main_headers = [
           'Label', 'Artist','Release','ISRC','ReleaseDate','Fanlink','MissingLinks'
    ]
    # Write the main data headers
    for col_num, header in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for release in Releases.objects.all():
        ws.append([
            release.Label,
            release.Artists,
            release.Title,
            release.UPC,
            release.ReleaseDate,
            release.FanlinkSent,
            release.MissingLinks
        ])

    ws.append([])       
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename=Releases_fanlinks.xlsx'

    return response



@csrf_exempt
def search_tracks(request):
    query = request.GET.get("query", "")
    if not query:
        return JsonResponse({"error": "Query parameter is required"}, status=400)

    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        #"q": f"{query} site:spotify.com",
        "q": f"{query} music artist",
        "key": settings.GOOGLE_API_KEY,  # Your Google API key
        "cx": settings.GOOGLE_CSE_ID,  # Your Custom Search Engine ID
        #"siteSearch": "youtube.com,spotify.com,boomplay.com,music.amazon.com,audiomack.com",
    }


    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        data = response.json()

        # Extract titles from the API response
        track_names = []
        if "items" in data:
            track_names = [item.get("title", "Unknown Track") for item in data["items"]]

        return JsonResponse({"tracks": track_names}, safe=False)
    except requests.exceptions.RequestException as e:
        return JsonResponse({"error": str(e)}, status=500)



class UploadVideoView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES['file']
        video = Video.objects.create(file=file_obj)

        # Extract just the filename (without "videos/")
        filename = os.path.basename(video.file.name)
        #folder_name = os.path.splitext(filename)[0]

        # Generate the correct streaming URL
        video_url = request.build_absolute_uri(
            reverse("stream_video", kwargs={"filename": filename})
        )

        return Response({
            "message": "Upload successful",
            "video_id": video.id,  # Include video_id in the response
            "video_url": video_url
        })


class UploadAccountSheetView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        # Get uploaded file
        file_obj = request.FILES['file_sheet']
        excel_data = file_obj.read()

        # Read Excel into DataFrame, preserve strings
        # First, get column names from the file
        temp_df = pd.read_excel(io.BytesIO(excel_data), skiprows=2, nrows=0)
        columns = temp_df.columns

        # Now read again with converters by column name
        df = pd.read_excel(
            io.BytesIO(excel_data),
            skiprows=2,
            converters={col: str for col in columns}
        )
        df = trim_at_first_empty_row(df)
        df.columns = df.columns.map(str)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df.replace([float('inf'), float('-inf')], '', inplace=True)
        df.fillna('', inplace=True)

        # Read date as string
        wb = load_workbook(io.BytesIO(excel_data), data_only=True)
        ws = wb.active
        raw_date = ws["B2"].value
        if isinstance(raw_date, datetime):
            date_value = raw_date.strftime("%b-%y")  # Example: Feb-25
        else:
            date_value = str(raw_date).strip()

        # Google Sheets auth
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        creds_path = os.path.join(BASE_DIR, "fanlink-440822-6316459498b3.json")
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
        client = gspread.authorize(creds)

        # Open spreadsheet and sheet
        spreadsheet = client.open_by_key("1J-nNwBetOsqu6EZk3-KsVBWJ_8sPgBc8AWnq9L-3yd0")
        #sheet = spreadsheet.worksheet("Sheet1")

        # Get current year
        current_year = str(datetime.now().year)

        # Create a new worksheet
        new_sheet_title = date_value+"-"+current_year
        sheet = spreadsheet.add_worksheet(title="Temporary", rows="100", cols="20")

        # Rename the new worksheet
        sheet.update_title(new_sheet_title)
        # Header values (B1 onwards)
        headers = [
            "SINGER", "TONENAME", "TONE_CD", "COUNT", "CHARGE", "REVENUE", "Label %",
            "Revenue (N)", "Label Share %", "Royalty Payable (N)", "AGENTS %",
            "Agents shar", "51 lex Records share", "Saheed@10%", "DANIEL 5%"
        ]
        # Write headers starting from cell B1
        start_col_index = 2  # Column B
        sheet.update_cell(1, start_col_index, headers[0])  # Write first header manually

        # Write remaining headers
        sheet.update(
            f'B1:{chr(65 + start_col_index + len(headers) - 1)}1',
            [headers]
        )

        from gspread.utils import rowcol_to_a1

        header_row = 1

        # Light gray background to A1
        dummy_fmt = CellFormat(backgroundColor=Color(0.95, 0.95, 0.95))
        format_cell_range(sheet, "A1", dummy_fmt)

        # Bold header
        bold_fmt = CellFormat(textFormat=TextFormat(bold=True))
        end_col_index = start_col_index + len(headers) - 1
        header_range = f"{rowcol_to_a1(header_row, start_col_index)}:{rowcol_to_a1(header_row, end_col_index)}"
        format_cell_range(sheet, header_range, bold_fmt)

        # Red "Agents shar"
        if "Agents shar" in headers:
            red_fmt = CellFormat(textFormat=TextFormat(bold=True, foregroundColor=Color(1, 0, 0)))
            red_col_index = headers.index("Agents shar") + start_col_index
            format_cell_range(sheet, rowcol_to_a1(header_row, red_col_index), red_fmt)

        # Blue "Saheed@10%"
        if "Saheed@10%" in headers:
            blue_fmt = CellFormat(textFormat=TextFormat(bold=True, foregroundColor=Color(0, 0, 1)))
            blue_col_index = headers.index("Saheed@10%") + start_col_index
            format_cell_range(sheet, rowcol_to_a1(header_row, blue_col_index), blue_fmt)



        # Determine where to append
        existing_data = sheet.get_all_values()
        start_row = 2
        sheet.update(f"B{start_row}", [[date_value]])

        # Update data rows
        table_start_row = start_row + 1
        sheet.update(f"B{table_start_row}", [df.columns.tolist()])
        max_cols = len(df.columns)

        # Format last column values to 2 decimal places
        last_col_index = df.shape[1] - 1  # 0-based index
        formatted_data = []
        for row in df.values.tolist():
            if len(row) > last_col_index:
                try:
                    row[last_col_index] = "{:.2f}".format(float(row[last_col_index]))
                except (ValueError, TypeError):
                    pass
            formatted_data.append(row + [''] * (max_cols - len(row)))

        sheet.update(f"B{table_start_row + 1}", formatted_data)
        table_end_row = table_start_row + len(df)


        # Add total if available
        xl = pd.read_excel(io.BytesIO(excel_data), header=None, dtype=str)

        total_value = xl.iloc[table_end_row + 1, 6] if xl.shape[0] > table_end_row + 1 else ""
        if total_value:
            try:
                total_value = "{:.2f}".format(float(total_value))
            except (ValueError, TypeError):
                pass
            sheet.update(f"G{table_end_row + 1}", [[total_value]])

        # 🎯 Formatting
        black = Color(0, 0, 0)

        border_format = CellFormat(
            borders=Borders(
                top=Border("SOLID", black),
                bottom=Border("SOLID", black),
                left=Border("SOLID", black),
                right=Border("SOLID", black),
            )
        )

        header_format = CellFormat(
            textFormat=TextFormat(bold=True, foregroundColor=black),
            borders=Borders(
                top=Border("SOLID", black),
                bottom=Border("SOLID", black),
                left=Border("SOLID", black),
                right=Border("SOLID", black),
            )
        )

        last_col_letter = colnum_to_letter(df.shape[1] + 1)  # +1 because we start from column B
        range_str = f"B{table_start_row}:{last_col_letter}{table_end_row}"

        # Apply formatting
        format_cell_range(sheet, range_str, border_format)
        format_cell_range(sheet, f"B{table_start_row}:{last_col_letter}{table_start_row}", header_format)

        # Apply 2-decimal formatting to last column
        value_range = f"{last_col_letter}{table_start_row + 1}:{last_col_letter}{table_end_row}"
        currency_format = CellFormat(
            numberFormat=NumberFormat(type='NUMBER', pattern='0.00')
        )
        format_cell_range(sheet, value_range, currency_format)



        # Compute and append calculated values to next column
        calculated_column_index = df.shape[1] + 1  # +1 because we're adding a new column
        calculated_column_letter_label = colnum_to_letter(calculated_column_index + 1)  # +1 for column B offset
        calculated_column_letter_revenue = colnum_to_letter(calculated_column_index + 2)
        calculated_column_letter_royalty_share = colnum_to_letter(calculated_column_index + 3)
        calculated_column_letter_royalty = colnum_to_letter(calculated_column_index + 4)

        calculated_column_letter_agent_percentage = colnum_to_letter(calculated_column_index + 5)
        calculated_column_letter_agent_share = colnum_to_letter(calculated_column_index + 6)
        calculated_column_letter_51lex = colnum_to_letter(calculated_column_index + 7)
        calculated_column_letter_saheed = colnum_to_letter(calculated_column_index + 8)
        calculated_column_letter_daniels = colnum_to_letter(calculated_column_index + 9)

        calculated_values = []
        rev_value = []
        royalty_share = []
        saheed_shares = []
        daniel_shares = []
        fifty1_shares = []
        royalty_pay = []
        agent_royalty_share = []
        agent_royalty_pay = []
        total_agent_share = 0
        total_royalty_share = 0
        total_revenue_share = 0
        second_col_index = df.shape[1] - 5
        for row in formatted_data[:-1]:
            try:
                last_value = float(row[last_col_index])
                rev = (26.32/100)*last_value
                revenue = round(rev, 2)  # Example: multiply by 1.1
                label_percentage = "26.32"
                tonename = str(row[second_col_index])
                label_share = 0
                agent_share = 0
                if tonename.strip() == "Gwo Gwo Ngwo":
                    label_share = 70
                    agent_share = 10
                elif tonename.strip() == "Gwo Gwo Ngwo Fast Jam Remix":
                    label_share = 30 
                    agent_share = 5
                elif tonename.strip() == "Gwo Gwo Ngwo Remix":
                    label_share = 30
                    agent_share = 5
                elif tonename.strip() == "Ka Esi Le Onye Isi Oche Gwo Gwo Ngwo":
                    label_share = 70 
                    agent_share = 10

                royalty_payable = (label_share/100)*revenue
                agent_share_payable = (agent_share/100)*royalty_payable
                fiftyOneLx = round(revenue - (agent_share_payable+royalty_payable),2)
                sah_share = round((10/100)*fiftyOneLx,2)
                dan_share = round((5/100)*fiftyOneLx,2)
                total_revenue_share = total_revenue_share + revenue
                total_agent_share = total_agent_share + agent_share_payable
                total_royalty_share = total_royalty_share + royalty_payable
    
            except (ValueError, TypeError):
                label_percentage = ''
                revenue = ""
                royalty_payable = ""
                label_share = ""
                agent_share_payable = ""
                fiftyOneLx = ""
                sah_share = ""
                dan_share = ""
            calculated_values.append([label_percentage])
            rev_value.append([revenue])
            royalty_share.append([label_share])
            royalty_pay.append([royalty_payable])
            agent_royalty_share.append([agent_share])
            agent_royalty_pay.append([agent_share_payable])
            saheed_shares.append([sah_share])
            daniel_shares.append([dan_share])
            fifty1_shares.append([fiftyOneLx])
            

        fiftyonelex_share = total_revenue_share - (total_agent_share+total_royalty_share)
        # Update the calculated column in Google Sheets 
        sheet.update(f"{calculated_column_letter_label}{table_start_row + 1}", calculated_values)
        sheet.update(f"{calculated_column_letter_revenue}{table_start_row + 1}", rev_value)
        total_row_number_rev = table_start_row + len(rev_value) + 1
        sheet.update(f"{calculated_column_letter_revenue}{total_row_number_rev}", [[format(round(total_revenue_share, 2), ",.2f")]])

        #royalty percentage
        sheet.update(f"{calculated_column_letter_royalty_share}{table_start_row + 1}", royalty_share)
        
        #royalty pay data
        sheet.update(f"{calculated_column_letter_royalty}{table_start_row + 1}", royalty_pay)
        total_row_number_royalty_pay = table_start_row + len(royalty_pay) + 1
        sheet.update(f"{calculated_column_letter_royalty}{total_row_number_royalty_pay}", [[format(round(total_royalty_share, 2), ",.2f")]])

        #agent percentage
        sheet.update(f"{calculated_column_letter_agent_percentage}{table_start_row + 1}", agent_royalty_share)
        

        #agent data appending
        sheet.update(f"{calculated_column_letter_agent_share}{table_start_row + 1}", agent_royalty_pay)
        total_row_number_agent_share = table_start_row + len(agent_royalty_pay) + 1
        sheet.update(f"{calculated_column_letter_agent_share}{total_row_number_agent_share}", [[format(round(total_agent_share, 2), ",.2f")]])
  
        #51 lex share
        sheet.update(f"{calculated_column_letter_51lex}{table_start_row + 1}", fifty1_shares)
        total_row_number_51lex = table_start_row + len(agent_royalty_pay) + 1
        sheet.update(f"{calculated_column_letter_51lex}{total_row_number_51lex}", [[format(round(fiftyonelex_share, 2), ",.2f")]])
        

        #Saheed share(10%)
        sheet.update(f"{calculated_column_letter_saheed}{table_start_row + 1}", saheed_shares)
        saheed_share = (10/100)*fiftyonelex_share
        total_row_number_saheed = table_start_row + len(agent_royalty_pay) + 1
        sheet.update(f"{calculated_column_letter_saheed}{total_row_number_saheed}", [[format(round(saheed_share, 2), ",.2f")]])
        

        #Daniel share(5%)
        sheet.update(f"{calculated_column_letter_daniels}{table_start_row + 1}", daniel_shares)
        daniel_share = (5/100)*fiftyonelex_share
        total_row_number_daniels = table_start_row + len(agent_royalty_pay) + 1
        sheet.update(f"{calculated_column_letter_daniels}{total_row_number_daniels}", [[format(round(daniel_share, 2), ",.2f")]])


                # --- Reload full sheet without header ---
        full_df = pd.read_excel(io.BytesIO(excel_data), header=None)

        # 1. Find the first blank row after first table
        blank_row_after_first_table = None
        for i in range(df.shape[0], full_df.shape[0]):
            row = full_df.iloc[i]
            if row.isnull().all() or row.astype(str).str.strip().eq("").all() :
                blank_row_after_first_table = i
                print("blank row detected after first table")
                break

        # 2. Start scanning for the second table after the blank row
        second_table_start = None
        second_table_end = None

        if blank_row_after_first_table is not None:
            for i in range(blank_row_after_first_table + 1, full_df.shape[0]):
                row = full_df.iloc[i]

                # Check if any cell in the row has meaningful data
                if row.notna().any() and row.astype(str).str.strip().any():
                    second_table_start = i
                    print(f"Second table starts at row {second_table_start + 1}")
                    break

        # Now find the end of the second table
        if second_table_start is not None:
            for j in range(second_table_start + 1, full_df.shape[0]):
                row = full_df.iloc[j]
                if not row.astype(str).str.strip().any():
                    second_table_end = j
                    break
            else:
                second_table_end = full_df.shape[0]

            # Slice out the second table
            second_table_df = full_df.iloc[second_table_start:second_table_end]

            # Drop fully empty columns (optional, in case there’s padding)
            second_table_df = second_table_df.dropna(how="all", axis=1)

            # 1. Find the next available row
            existing_values = sheet.get_all_values()
            next_available_row = len(existing_values) + 4  # +4 for blank space after first table

            # 2. Convert DataFrame to list of lists
            second_table_values = second_table_df.values.tolist()

            # 3. Update values to Google Sheet in columns C and D
            cell_range = f"C{next_available_row}"
            sheet.update(cell_range, second_table_values)

            # 4. Format second column (figures) as number — it's now column D
            end_row = next_available_row + len(second_table_values) - 1
            format_cell_range(
                sheet,
                f"D{next_available_row}:D{end_row}",
                CellFormat(
                    numberFormat={
                        "type": "NUMBER",
                        "pattern": "#,##0.00"
                    }
                )
            )

            # 5. Apply borders to the 2-column second table (columns C and D)
            format_cell_range(
                sheet,
                f"C{next_available_row}:D{end_row}",
                CellFormat(
                    borders=Borders(
                        top=Border("SOLID", Color(0, 0, 0)),
                        bottom=Border("SOLID", Color(0, 0, 0)),
                        left=Border("SOLID", Color(0, 0, 0)),
                        right=Border("SOLID", Color(0, 0, 0)),
                    )
                )
            )


            # 1. Perform last calculation for premier records and others
            extra_rows = [
                ["PREMIER RECORDS LTD SHARE PAYABLE", total_royalty_share],
                ["AGENT", total_agent_share],
                ["51 LEX RECORDS", fiftyonelex_share],
                ["SAHEED", saheed_share],
                ["DANIEL", daniel_share],
            ]

            # 2. Determine starting row for extra data (just after second table)
            extra_row_start = end_row + 1
            extra_row_end = extra_row_start + len(extra_rows) - 1

            # 3. Update values into Google Sheet (columns C and D)
            sheet.update(f"C{extra_row_start}:D{extra_row_end}", extra_rows)

            # 4. Format values in column D as numbers with thousands comma
            format_cell_range(
                sheet,
                f"D{extra_row_start}:D{extra_row_end}",
                CellFormat(
                    numberFormat={
                        "type": "NUMBER",
                        "pattern": "#,##0.00"
                    }
                )
            )

            # 5. Apply red text formatting to all inserted rows
            format_cell_range(
                sheet,
                f"C{extra_row_start}:D{extra_row_end}",
                CellFormat(
                    textFormat=TextFormat(
                        foregroundColor=Color(1, 0, 0),  # Red text
                        bold=True
                    )
                )
            )

     


       # Extract tables
        first_table_start_col = 1  # column B
        second_table_start_col = 2  # column C

        # Find where each table ends by checking for the first row where the entire table is NaN
        def get_table_rows(df, start_col):
            data_rows = []
            for idx in range(2, len(df)):  # skip row 0 and 1 (title + date)
                row = df.iloc[idx]
                values = row[start_col:]
                if values.isnull().all():
                    break
                data_rows.append(values.tolist())
            return data_row

        first_table_df = df

        # Entity-specific percentages
        entity_percentages = {
            "Saheed": 0,
            "Daniel": 0,
            "AgentMO": 0,
            "PREMIER RECORDS":0
        }

        thin = Side(border_style="thin", color="000000")
        thin_border = XLBorder(left=thin, right=thin, top=thin, bottom=thin)

        # Prepare ZIP
        zip_buffer = BytesIO()

        with ZipFile(zip_buffer, 'w') as zipf:
            for entity, percent in entity_percentages.items():
                # Make copy of original first table
                df_raw = first_table_df.copy()

                # Identify the last column
                last_col_name = df_raw.columns[-1]

                # Avoid total row for calculation
                df_calc = df_raw.iloc[:-1].copy()

               
                 # Add extra columns
                df_calc[last_col_name] = pd.to_numeric(df_calc[last_col_name], errors='coerce')
                df_calc["Label %"] = 26.32
                df_calc["Revenue (N)"] = ( (26.32 / 100)*df_calc[last_col_name]).round(2)
                # Initialize new columns base on agent
                if entity=="PREMIER RECORDS":
                    df_calc["Label share %"] = 0
                    df_calc["Royalty Payable"] = 0
                elif entity=="AgentMO":
                    df_calc["Label share %"] = 0
                    df_calc["Royalty Payable"] = 0
                    df_calc["AGENTS %"] = 0
                    df_calc["Agents Share"] = 0
                elif entity=="Daniel":
                    df_calc["Label share %"] = 0
                    df_calc["Royalty Payable"] = 0
                    df_calc["51 lex Records share"] = ''
                    df_calc["Daniel 5%"] = 0
                elif entity=="Saheed":
                    df_calc["Label share %"] = 0
                    df_calc["Royalty Payable"] = 0
                    df_calc["51 lex Records share"] = ''
                    df_calc["Saheed 10%"] = 0

                

                second_col_name = df_raw.columns[1]
                # Assign values row by row
                for idx, row in df_calc.iterrows():
                    song_title = row[second_col_name]
                    revenue = row["Revenue (N)"]
                    if song_title == "Gwo Gwo Ngwo" or song_title == "Ka Esi Le Onye Isi Oche Gwo Gwo Ngwo":
                        label_share = 70
                        agent_percent = 10
                    elif song_title in ["Gwo Gwo Ngwo Fast Jam Remix", "Gwo Gwo Ngwo Remix"]:
                        label_share = 30
                        agent_percent = 5
                    else:
                        label_share = 0
                        agent_percent = 0

                    royalty = (label_share / 100) * revenue
                    agent_share = (agent_percent / 100) * royalty

                    if entity=="PREMIER RECORDS":
                        df_calc.at[idx, "Label share %"] = int(label_share)
                        df_calc.at[idx, "Royalty Payable"] = round(royalty, 2)
                    elif entity=="AgentMO":
                        df_calc.at[idx, "Label share %"] = int(label_share)
                        df_calc.at[idx, "Royalty Payable"] = round(royalty, 2)
                        df_calc.at[idx, "AGENTS %"] = int(agent_percent)
                        df_calc.at[idx, "Agents Share"] = round(agent_share, 2)
                    elif entity=="Daniel":
                        df_calc.at[idx, "Label share %"] = int(label_share)
                        df_calc.at[idx, "Royalty Payable"] = round(royalty, 2)
                        fifty1 = revenue - (agent_share+royalty)
                        df_calc.at[idx, "Daniel 5%"] = round(fifty1 * 5 / 100, 2)
                    elif entity=="Saheed":
                        df_calc.at[idx, "Label share %"] = int(label_share)
                        df_calc.at[idx, "Royalty Payable"] = round(royalty, 2)
                        fifty1 = revenue - (agent_share+royalty) 
                        df_calc.at[idx, "Saheed 10%"] = round(fifty1 * 10 / 100, 2)
                    
                    
                    
                # Concatenate total row back
                total_row = df_raw.iloc[[-1]].copy()  
                total_row["Label %"] = ''
                total_row["Revenue (N)"] = (df_calc["Revenue (N)"].sum()).round(2)

                if entity=="PREMIER RECORDS":
                    total_row["Label share %"] = ''
                    total_row["Royalty Payable"] = (df_calc["Royalty Payable"].sum()).round(2)
                elif entity=="AgentMO":
                    total_row["Label share %"] = ''
                    total_row["Royalty Payable"] = (df_calc["Royalty Payable"].sum()).round(2)
                    total_row["AGENTS %"] = ''
                    total_row["Agents Share"] = (df_calc["Agents Share"].sum()).round(2)
                elif entity=="Daniel":
                    total_row["Label share %"] = ''
                    total_row["Royalty Payable"] = (df_calc["Royalty Payable"].sum()).round(2)
                    total_row["51 lex Records share"] = round(fiftyonelex_share,2)
                    total_row["Daniel 5%"] = (df_calc["Daniel 5%"].sum()).round(2)
                elif entity=="Saheed":
                    total_row["Label share %"] = ''
                    total_row["Royalty Payable"] = (df_calc["Royalty Payable"].sum()).round(2)
                    total_row["51 lex Records share"] = round(fiftyonelex_share,2)
                    total_row["Saheed 10%"] = (df_calc["Saheed 10%"].sum()).round(2)

                    
                
                final_table_df = pd.concat([df_calc, total_row], ignore_index=True)

                # Create Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Account Report"

                # Set date
                ws['B2'] = date_value

                # Append second table first (no headers)
                ws.append([])
                second_table_start_row = ws.max_row + 2
                for r in dataframe_to_rows(second_table_df, index=False, header=False):
                    ws.append(['', ''] + r)
                second_table_end_row = ws.max_row
                second_table_cols = len(second_table_df.columns)
                if entity=="PREMIER RECORDS":
                    ws.append(['','','PREMIER RECORDS LTD SHARE PAYABLE',round(total_royalty_share,2)])
                elif entity=="AgentMO":
                    ws.append(['','','PREMIER RECORDS LTD SHARE PAYABLE',round(total_royalty_share,2)])
                    ws.append(['','','AGENTMO SHARE',round(total_agent_share,2)])
                elif entity=="Daniel":
                    ws.append(['','','PREMIER RECORDS LTD SHARE PAYABLE',round(total_royalty_share,2)])
                    ws.append(['','','51 LEX RECORDS SHARE', round(fiftyonelex_share,2)])
                    ws.append(['','','DANIEL 10%', round(daniel_share,2)])
                elif entity=="Saheed":
                    ws.append(['','','PREMIER RECORDS LTD SHARE PAYABLE',round(total_royalty_share,2)])
                    ws.append(['','','51 LEX RECORDS SHARE', round(fiftyonelex_share,2)])
                    ws.append(['','','SAHEED 5%', round(saheed_share,2)])

                    
                # Append first table (with headers and extra columns)
                ws.append([])
                ws.append([])
                first_table_start_row = ws.max_row + 3
                for r in dataframe_to_rows(final_table_df, index=False, header=True):
                    ws.append([''] + r)
                first_table_end_row = ws.max_row
                first_table_cols = len(final_table_df.columns)

                # Apply border to second table (now first)
                for row in ws.iter_rows(min_row=second_table_start_row, max_row=second_table_end_row, min_col=3, max_col=2 + second_table_cols):
                    for cell in row:
                        cell.border = thin_border

                # Apply border and formatting to first table (now second)
                for row_idx, row in enumerate(ws.iter_rows(min_row=first_table_start_row, max_row=first_table_end_row, min_col=2, max_col=1 + first_table_cols), start=0):
                    for col_idx, cell in enumerate(row):
                        cell.border = thin_border
                        if row_idx == 0:  # Header row
                            cell.font = Font(bold=True)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'

                # Save to memory
                file_buffer = BytesIO()
                wb.save(file_buffer)
                file_buffer.seek(0)

                # Add to ZIP
                zipf.writestr(f"{entity}-account_statement.xlsx", file_buffer.read())

        # Finalize ZIP
        zip_buffer.seek(0)
        # Return ZIP as downloadable response
        response = FileResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="Agents-Account_Reports.zip"'
        return response



        # return Response({
        #     "message": "Uploaded and appended successfully"
        # })


def trim_at_first_empty_row(dataframe):
    for idx, row in dataframe.iterrows():
        if row.isnull().all() or (row.astype(str).str.strip() == '').all():
            return dataframe.iloc[:idx]
    return dataframe


    
def colnum_to_letter(n):
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


    

def serve_video(request, filename):
    """Serve video with proper headers for seeking support."""
    file_path = os.path.join(settings.MEDIA_ROOT, "videos", filename)  # Adjusted path

    if not os.path.exists(file_path):
        return HttpResponse(status=404)

    # Serve the file with support for byte-range requests
    response = FileResponse(open(file_path, "rb"), content_type="video/mp4")
    response["Accept-Ranges"] = "bytes"
    
    return response


def add_watermark_to_video(input_video, output_video, watermark_image):
    """
    Adds a watermark to the given video using FFmpeg.
    """
    try:
        command = [
            "ffmpeg",
            "-i", input_video,  # Input video
            "-i", watermark_image,  # Watermark image
            "-filter_complex", "overlay=W-w-10:H-h-10",  # Position watermark at bottom-right
            "-codec:a", "copy",  # Keep original audio
            output_video
        ]
        subprocess.run(command, check=True)  # Run FFmpeg command
        return output_video  # Return watermarked video path
    except subprocess.CalledProcessError as e:
        print(f"Error adding watermark: {e}")
        return None  # Return None if an error occurs




@csrf_exempt
@api_view(["POST"])
@parser_classes([MultiPartParser])
def trim_video(request):
    try:
        video_id = request.data.get("video_id")
        start_time = float(request.data.get("start_time", 0))
        end_time = float(request.data.get("end_time", 0))
        watermark_file = request.FILES.get("watermark_image")
        watermark_temp_path = None
        

        if not video_id or start_time < 0 or end_time <= start_time:
            return JsonResponse({"error": "Invalid input data"}, status=400)

        # Get video from DB
        try:
            video = Video.objects.get(id=video_id)
        except Video.DoesNotExist:
            return JsonResponse({"error": "Video not found"}, status=404)

        input_video = os.path.join(settings.MEDIA_ROOT, str(video.file))
        video_name = os.path.splitext(os.path.basename(input_video))[0]
        trimmed_filename = f"{video_name}_trimmed.mp4"
        trimmed_path = os.path.join(settings.MEDIA_ROOT, "trimmed", trimmed_filename)
        watermarked_path = os.path.join(settings.MEDIA_ROOT, "trimmed", f"watermarked_{trimmed_filename}")
        

        # Remove old versions if they exist
        if os.path.exists(trimmed_path):
            os.remove(trimmed_path)
        if os.path.exists(watermarked_path):
            os.remove(watermarked_path)

        # Trim the video using FFmpeg
        subprocess.run([
            "ffmpeg", "-i", input_video, "-ss", str(start_time), "-to", str(end_time),
            "-c", "copy", trimmed_path
        ], check=True)

        # Handle uploaded watermark image
        if watermark_file:
            filename_base = os.path.splitext(watermark_file.name)[0]
            watermark_temp_path = os.path.join(settings.MEDIA_ROOT, f"resized_{filename_base}.jpg")
            img = Image.open(watermark_file)
            if img.mode in ("RGBA", "P"):
               img = img.convert("RGB")  # Convert to RGB to remove alpha

            img = img.resize((120, 80))
            img.save(watermark_temp_path, "JPEG")

            # Watermark the trimmed video
            subprocess.run([
                "ffmpeg", "-i", trimmed_path, "-i", watermark_temp_path,
                "-filter_complex", "overlay=W-w-10:H-h-10", "-codec:a", "copy", watermarked_path
            ], check=True)

            os.remove(trimmed_path)  # Delete the trimmed version
            os.remove(watermark_temp_path)  # Delete temp watermark image
        else:
            watermark_image = os.path.join(settings.MEDIA_ROOT, "Africha_Entertainment.png")
            watermarked_video = add_watermark_to_video(trimmed_path, watermarked_path, watermark_image)
            if watermarked_video:
               os.remove(trimmed_path)  # Remove original trimmed video

        # Return final video URL
        trimmed_video_url = request.build_absolute_uri(settings.MEDIA_URL + "trimmed/" + os.path.basename(watermarked_path))

        return JsonResponse({
            "message": "Video trimmed and watermarked successfully",
            "trimmed_video_url": trimmed_video_url
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def download_trimmed_video(request, filename):
    file_path = os.path.join(settings.MEDIA_ROOT, "trimmed", filename)
    if os.path.exists(file_path):
        return FileResponse(open(file_path, "rb"), content_type="video/mp4")
    else:
        raise Http404("Video not found")

@csrf_exempt
@api_view(["POST"])
@parser_classes([MultiPartParser])
def split_video(request):
    video_id = request.data.get("video_id")
    duration = int(request.data.get("duration", 15))
    watermark_file = request.FILES.get("watermark_image")
    watermark_temp_path = ""
    
        

    try:
        video = Video.objects.get(id=video_id)
        video_path = os.path.join(settings.MEDIA_ROOT, str(video.file))
       
        if watermark_file: 
            # Save watermark temporarily
            filename_base = os.path.splitext(watermark_file.name)[0]
            watermark_temp_path = os.path.join(settings.MEDIA_ROOT, f"resized_{filename_base}.jpg")
            img = Image.open(watermark_file)
            if img.mode in ("RGBA", "P"):
               img = img.convert("RGB")  # Convert to RGB to remove alpha

            img = img.resize((120, 80))
            img.save(watermark_temp_path, "JPEG")

        # Create output folder
        folder_name = os.path.splitext(os.path.basename(video_path))[0]
        output_folder = os.path.join(settings.MEDIA_ROOT, "splitted_videos", folder_name)
        if os.path.exists(output_folder):
            random_chars = ''.join(random.choices(string.ascii_letters + string.digits, k=3))
            folder_name += random_chars
            output_folder = os.path.join(settings.MEDIA_ROOT, "splitted_videos", folder_name)

        os.makedirs(output_folder, exist_ok=True)

        # Get total duration
        result = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", video_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        total_duration = float(result.stdout.strip())

        # Split and watermark
        part = 1
        start_time = 0
        while start_time < total_duration:
            temp_file = os.path.join(output_folder, f"temp_part{part}.mp4")
            output_file = os.path.join(output_folder, f"{folder_name}_part{part}.mp4")

            # Split the video
            subprocess.run([
                "ffmpeg", "-i", video_path, "-ss", str(start_time), "-t", str(duration),
                "-c", "copy", temp_file
            ], check=True)
            
            if watermark_file :
                # Add watermark if uploaded
                subprocess.run([
                    "ffmpeg", "-i", temp_file, "-i", watermark_temp_path,
                    "-filter_complex", "overlay=W-w-10:H-h-10", "-codec:a", "copy", output_file
                ], check=True)
                os.remove(temp_file)
            else : 
                # add water mark if not uploaded
                watermark_image = os.path.join(settings.MEDIA_ROOT, "Africha_Entertainment.png")
                watermarked_video = add_watermark_to_video(temp_file, output_file, watermark_image)
                if watermarked_video:
                  os.remove(temp_file)  # Delete temp file after watermarking

            start_time += duration
            part += 1

        # Clean up watermark image
        if watermark_file :
            if os.path.exists(watermark_temp_path):
                os.remove(watermark_temp_path)

        video.splitted = "Yes"
        video.save()

        return JsonResponse({
            "message": "Video split and watermarked successfully",
            "download_url": folder_name
        })

    except Video.DoesNotExist:
        return JsonResponse({"error": f"Video not found: id = {video_id}"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def download_split_folder(request, folder_name):
    folder_path = os.path.join(settings.MEDIA_ROOT, "splitted_videos", folder_name)
    zip_path = f"{folder_path}.zip"
    
    if not os.path.exists(folder_path):
        return JsonResponse({"error": "Folder not found"}, status=404)
    
    # Zip the folder
    shutil.make_archive(folder_path, 'zip', folder_path)
    return FileResponse(open(zip_path, "rb"), as_attachment=True, filename=f"{folder_name}.zip")


@csrf_exempt
def get_uploaded_videos(request):
    videos = Video.objects.all()
    video_list = [
        {
            "id": video.id,
            "name": os.path.basename(video.file.name),  # Extracts video file name
            "file_url": request.build_absolute_uri(settings.MEDIA_URL + str(video.file)),
            "uploaded_at": video.uploaded_at.strftime("%Y-%m-%d %H:%M:%S"),
            "splitted": video.splitted
        }
        for video in videos
    ]
    return JsonResponse({"videos": video_list}, safe=False)



@csrf_exempt
def delete_video(request, video_id):
    if request.method == "DELETE":
        video = get_object_or_404(Video, id=video_id)

        # Delete the video file from storage
        if video.file:
            trimmed_filename = f"{video_id}_trimmed.mp4"
            splitted_folder = os.path.splitext(os.path.basename(video.file.name))[0]
            zipfile = f"{splitted_folder}.zip"
            split_folder_path = os.path.join(settings.MEDIA_ROOT, "splitted_videos", splitted_folder)
            zipfile_path = os.path.join(settings.MEDIA_ROOT, "splitted_videos", zipfile)
            trim_video_path = os.path.join(settings.MEDIA_ROOT, "trimmed", trimmed_filename)
            video_path = video.file.path  # Absolute path of the video file
            if os.path.exists(video_path):
                os.remove(video_path)  
            if os.path.exists(zipfile_path):
                os.remove(zipfile_path)
            if os.path.exists(trim_video_path):
                os.remove(trim_video_path)  
            if os.path.exists(split_folder_path):
                shutil.rmtree(split_folder_path)
                 

        # Delete from database
        video.delete()
        
        return JsonResponse({"message": "Video deleted successfully"}, status=200)

    return JsonResponse({"error": "Invalid request"}, status=400)
